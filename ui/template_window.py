import os

import requests
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QSplitter, QLabel, QPushButton, QListWidget, QTableWidget, \
    QHBoxLayout, QListWidgetItem, QFileDialog, QDialog, QMessageBox, QTableWidgetItem, QProgressDialog, QInputDialog

from openpyxl import load_workbook


class TemplateTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.zoom_level = 100
        self.default_font_size = self.font().pointSize()

    def wheelEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_level += 10
            else:
                self.zoom_level -= 10
                if self.zoom_level < 10:
                    self.zoom_level = 10

            scale_factor = self.zoom_level / 100
            self.horizontalHeader().setDefaultSectionSize(int(100 * scale_factor))
            self.verticalHeader().setDefaultSectionSize(int(30 * scale_factor))

            new_font_size = int(self.default_font_size * scale_factor)
            self.update_table_font(new_font_size)

            event.accept()
        else:
            super().wheelEvent(event)


    def update_table_font(self, font_size):
        """Update the font size of all cells in the table while preserving existing font styles."""
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    current_font = item.font()
                    current_font.setPointSize(font_size * 2)
                    item.setFont(current_font)


class Step2(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.backend_url = "http://2.56.178.70:8000"
        self.main_window = main_window

        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        splitter = QSplitter(Qt.Orientation.Horizontal)

        self.label = QLabel("Шаг 2: Загрузить Шаблон")

        self.load_local_button = QPushButton("Загрузить Шаблон с Компьютера")
        self.load_local_button.clicked.connect(self.load_local_file)

        self.load_cloud_button = QPushButton("Загрузить Шаблон с Сервера")
        self.load_cloud_button.clicked.connect(self.load_cloud_file)

        self.raw_data_label = QLabel("Столбцы из файла данных:")
        self.raw_data_list = QListWidget()

        self.template_table = TemplateTable()
        self.template_table.itemChanged.connect(self.on_template_changed)

        self.button_layout = QHBoxLayout()
        self.save_button = QPushButton("Сохранить изменения")
        self.save_button.clicked.connect(self.save_changes)
        self.save_button.setEnabled(False)

        self.generate_button = QPushButton("Генерация файлов")
        self.generate_button.clicked.connect(self.generate_files)
        self.generate_button.setEnabled(False)

        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.generate_button)

        left_panel = QWidget()
        left_layout.addWidget(self.raw_data_label)
        left_layout.addWidget(self.raw_data_list)
        left_panel.setLayout(left_layout)

        right_panel = QWidget()
        right_layout.addWidget(self.load_local_button)
        right_layout.addWidget(self.load_cloud_button)
        right_layout.addWidget(self.template_table)
        right_layout.addLayout(self.button_layout)
        right_panel.setLayout(right_layout)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 5)

        main_layout = QVBoxLayout()
        main_layout.addWidget(splitter)

        self.setLayout(main_layout)

        self.template_path = ""
        self.saved_template_path = ""
        self.merged_ranges = []
        self.workbook = None
        self.sheet = None
        self.changes = {}
        self.placeholders = {}

    def update_raw_data_columns(self, columns):
        self.raw_data_list.clear()
        for idx, column_name in enumerate(columns, start=1):
            item = QListWidgetItem(f"{idx}. {column_name}")
            self.raw_data_list.addItem(item)

    def on_template_changed(self, item):
        if item and self.template_path:
            self.save_button.setEnabled(True)
            if self.sheet:
                row = item.row()
                col = item.column()

                self.changes[(row + 1, col + 1)] = item.text()

    def detect_placeholders(self, cell_value, row, col):
        if cell_value and isinstance(cell_value, str):
            import re
            matches = re.findall(r"\{(\d+)\}", cell_value)
            if matches:
                self.placeholders[(row, col)] = {
                    'template': cell_value,
                    'column_ids': [int(m) - 1 for m in matches]
                }

    def load_local_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выбрать Шаблон", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.template_path = file_path
            self.display_template()
            self.save_button.setEnabled(True)

    def load_cloud_file(self):
        try:
            response = requests.get(f"{self.backend_url}/templates/")
            response.raise_for_status()
            templates = response.json()

            dialog = QDialog(self)
            dialog.setWindowTitle("Выбрать Шаблон с Сервера")
            layout = QVBoxLayout()
            list_widget = QListWidget()
            for template in templates:
                item = QListWidgetItem(template['name'])
                item.setData(Qt.ItemDataRole.UserRole, template['id'])
                list_widget.addItem(item)
            layout.addWidget(list_widget)

            load_button = QPushButton("Выбрать")
            load_button.clicked.connect(lambda: self.fetch_template_from_cloud(list_widget, dialog))
            layout.addWidget(load_button)

            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при получнии Шаблона из сервера: {str(e)}")

    def fetch_template_from_cloud(self, list_widget, dialog):
        selected_item = list_widget.currentItem()
        if selected_item:
            template_name = selected_item.text()
            template_id = selected_item.data(Qt.ItemDataRole.UserRole)
            try:
                response = requests.get(f"{self.backend_url}/templates/{template_id}/")
                response.raise_for_status()

                with open(template_name, "wb", encoding="utf-8") as f:
                    f.write(response.content)

                self.template_path = template_name
                self.display_template()
                dialog.accept()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при получнии Шаблона из сервера: {str(e)}")

    def display_template(self):
        if self.template_path:
            try:
                self.workbook = load_workbook(self.template_path)
                self.sheet = self.workbook.active

                self.template_table.setRowCount(self.sheet.max_row)
                self.template_table.setColumnCount(self.sheet.max_column)

                self.merged_ranges = []
                self.placeholders = {}
                self.changes = {}

                for row_idx, row in enumerate(self.sheet.iter_rows()):
                    for col_idx, cell in enumerate(row):
                        value = cell.value
                        item = QTableWidgetItem(str(value) if value is not None else "")

                        if cell.font:
                            font = QFont()
                            if cell.font.name:
                                font.setFamily(cell.font.name)
                            if cell.font.size:
                                font.setPointSize(int(cell.font.size))
                            font.setBold(cell.font.bold or False)
                            font.setItalic(cell.font.italic or False)
                            item.setFont(font)

                        if cell.alignment and cell.alignment.horizontal:
                            if cell.alignment.horizontal == "center":
                                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            elif cell.alignment.horizontal == "left":
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft)
                            elif cell.alignment.horizontal == "right":
                                item.setTextAlignment(Qt.AlignmentFlag.AlignRight)

                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            rgb = cell.fill.start_color.rgb
                            if isinstance(rgb, str) and len(rgb) >= 6:
                                color = QColor(f"#f0f0f0")
                                item.setBackground(color)

                        self.template_table.setItem(row_idx, col_idx, item)

                for merged_range in self.sheet.merged_cells.ranges:
                    min_row = merged_range.min_row - 1
                    min_col = merged_range.min_col - 1
                    row_span = merged_range.max_row - merged_range.min_row + 1
                    col_span = merged_range.max_col - merged_range.min_col + 1
                    self.template_table.setSpan(min_row, min_col, row_span, col_span)
                    self.merged_ranges.append((min_row, min_col, row_span, col_span))

                self.template_table.resizeColumnsToContents()
                self.template_table.resizeRowsToContents()

                if self.placeholders:
                    QMessageBox.information(
                        self,
                        "Placeholders Detected",
                        f"Found {len(self.placeholders)} cells with placeholders in the template."
                    )

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при получении шаблона: {str(e)}")

    def save_changes(self):
        self.save_changes_local()

        reply = QMessageBox.question(
            self,
            "Обновление Шаблона",
            "Обновить Шаблон на Сервере?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                with open(self.saved_template_path, "rb", encoding="utf-8") as f:
                    files = {"file": f}
                    response = requests.post(f"{self.backend_url}/templates/update/", files=files)
                    response.raise_for_status()
                    QMessageBox.information(self, "Отлично", "Новый шаблон загружен на сервер.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке шаблона на сервер: {str(e)}")

    def save_changes_local(self):
        if not self.sheet:
            QMessageBox.warning(self, "Предупреждение", "Сначала нужно указать файл Шаблона.")
            return

        try:
            for merged_range in self.sheet.merged_cells.ranges.copy():
                self.sheet.unmerge_cells(str(merged_range))

            for row in range(self.template_table.rowCount()):
                for col in range(self.template_table.columnCount()):
                    item = self.template_table.item(row, col)
                    if item:
                        cell = self.sheet.cell(row=row + 1, column=col + 1)
                        cell.value = self.changes.get((row + 1, col + 1), item.text())

                        self.detect_placeholders(cell.value, row + 1, col + 1)

            for min_row, min_col, row_span, col_span in self.merged_ranges:
                start_row = min_row + 1
                start_col = min_col + 1
                end_row = start_row + row_span - 1
                end_col = start_col + col_span - 1

                value = self.sheet.cell(row=start_row, column=start_col).value

                self.sheet.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=end_col
                )

                self.sheet.cell(row=start_row, column=start_col).value = value

            if self.template_path.endswith('_marked.xlsx'):
                self.saved_template_path = self.template_path
            else:
                self.saved_template_path = self.template_path.replace('.xlsx', '_marked.xlsx')
            self.workbook.save(self.saved_template_path)
            self.changes = {}
            QMessageBox.information(self, "Отлично", f"Шаблон сохранен как: {self.saved_template_path}")
            self.generate_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении шаблона: {str(e)}")


    def generate_files(self):
        if not self.saved_template_path or self.main_window.step1.data is None:
            QMessageBox.warning(self, "Предупреждение", "Удостовертесь что шаблон и файл с данными загружен.")
            return

        chosen_dir = QFileDialog.getExistingDirectory(self, "Выбрать файл сохранения")
        if not chosen_dir:
            return

        file_type, ok = QInputDialog.getItem(
            self,
            "Выбрать тип файла",
            "Выберите формат:",
            ["Excel", "PDF"],
            0,
            False
        )
        if not ok:
            return

        name_pattern, ok = QInputDialog.getText(
            self,
            "Паттерн Названия Файла",
            "Укажите паттерн для названия файла (используйте {1}, {9}):",
            text="{2}_{1}_gen"
        )
        if not ok:
            return

        try:
            raw_data = self.main_window.step1.data

            output_dir_xlsx = os.path.join(chosen_dir, os.path.basename(self.template_path).replace('.xlsx', '_filled_excel'))
            output_dir_pdf = os.path.join(chosen_dir, os.path.basename(self.template_path).replace('.xlsx', '_filled_pdf'))
            os.makedirs(output_dir_xlsx, exist_ok=True)
            os.makedirs(output_dir_pdf, exist_ok=True)

            total_files = len(raw_data)
            progress = QProgressDialog("Генерация файлов...", None, 0, total_files, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)

            for idx, row in enumerate(raw_data.itertuples(index=False), start=1):
                file_name = name_pattern
                for col_idx, col_value in enumerate(row, start=1):  # Adjusting for 1-based index
                    file_name = file_name.replace(f"{{{col_idx}}}", str(col_value))

                new_wb = load_workbook(self.saved_template_path)
                new_sheet = new_wb.active

                for (row_idx, col_idx), placeholder_info in self.placeholders.items():
                    template_text = placeholder_info['template']
                    for col_id in placeholder_info['column_ids']:
                        if col_id < len(raw_data.columns):
                            value = str(row[col_id])
                            template_text = template_text.replace(f"{{{col_id + 1}}}", value)

                    new_sheet.cell(row=row_idx, column=col_idx).value = template_text

                output_path = os.path.join(output_dir_xlsx, file_name) + ".xlsx"
                output_dir = output_dir_xlsx
                new_wb.save(output_path)
                new_wb.close()

                if file_type == "PDF":
                    pdf_path = os.path.join(output_dir_pdf, file_name) + ".pdf"
                    self.convert_excel_to_pdf_windows(output_path, pdf_path)
                    output_dir = output_dir_pdf

                progress.setValue(idx)

            QMessageBox.information(
                self,
                "Отлично",
                f"Сгенерировано {total_files} файлов в папке: {output_dir}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при генерации файлов: {str(e)}")
            raise e


    @staticmethod
    def convert_excel_to_pdf_windows(input_excel_path, output_pdf_path):

        import os

        from win32com.client import Dispatch, constants
        """
            Converts an Excel file to a PDF file on Windows using win32com.client.
            Args:        input_excel_path (str): Path to the input Excel file.
                output_pdf_path (str): Path to the output PDF file.
            Raises:        Exception: If there is an error during the conversion.
            """
        excel = None
        try:
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(input_excel_path)
            workbook.ExportAsFixedFormat(0, output_pdf_path)

        except Exception as e:
            raise Exception(f"Ошибка при конвертации excel на pdf: {str(e)}")
        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()

        if not os.path.exists(output_pdf_path):
            raise Exception("Ошибка. PDF файл не создан")


